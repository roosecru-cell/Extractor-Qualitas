import streamlit as st
import pdfplumber
import re
import os
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from itertools import groupby

st.set_page_config(
    page_title="Extractor Valuaciones – Quálitas",
    page_icon="🔧",
    layout="centered"
)

st.markdown("""
<style>
    .block-container { padding-top: 2rem; }
    h1 { color: #1a3a5c; }
    .stButton>button {
        background-color: #1a3a5c;
        color: white;
        border-radius: 6px;
        padding: 0.5rem 2rem;
        font-weight: 600;
    }
    .stButton>button:hover { background-color: #c8392b; }
    .result-box {
        background: white;
        border-radius: 8px;
        padding: 1rem 1.5rem;
        margin-top: 1rem;
        border-left: 4px solid #1a3a5c;
        box-shadow: 0 2px 6px rgba(0,0,0,0.08);
    }
</style>
""", unsafe_allow_html=True)

st.title("🔧 Extractor de Valuaciones Quálitas")
st.caption("Refacciones · Pintura · Hojalatería · Mecánica  |  Sube uno o varios PDFs")

# ── Conceptos de Mecánica ─────────────────────────────────────────────────────
# Cuando aparecen en Hojalatería se reclasifican como Mecánica

CONCEPTOS_MECANICA = {
    "RIN DL.D.:D+M",
    "RIN DL.I.:D+M",
    "RIN DEL.D. REPARAR",
    "BRAZO SUSPENS.DL.I.INF.:D+M",
    "LLANTA DL.D.:D+M",
    "LLANTA DL.I.:D+M",
    "LLANTA DL.I.:D+M Y BALANCEAR",
    "RADIADOR:D+M",
    "RIN DEL.I. REPARAR",
    "03 17 00 LIQUIDO REFRIG.:VAC-LLENAR",
    "05 19 00 RIN DL.D.:D+M",
    "38 17 70 LIQUIDO REFRIG.:VACIAR-RELLENAR",
    "50 00 ZAX FUNCION GFS/AJUSTES",
    "AIRE ACONDIC.:VAC-LLENAR",
    "ALINEACION 260415174154772",
    "ALINEACION 260505175144873",
    "ALINEACION 260526102213567",
    "ALINEACION Y BALANCE 260514100700748",
    "ALINEACION Y BALANCE 260519161659496",
    "AMORTIG.DL.D.:D+M",
    "AMORTIG.DL.D.:DESPIEZ-ENSAMB.(DESMONT.)",
    "AMORTIG.DL.I.:DESPIEZ-ENSAMBLAR(DESMONT)",
    "AMORTIG.DL.I./D.:DESP-ENSAMB.(DESMONT.)",
    "AMORTIG.DL.I.CPL.:D+M",
    "AMORTIG.FACIA DL.:D+M",
    "ANTICONGELANTE 260505174818878",
    "ANTICONGELANTE 260512135935885",
    "ARNES 260519083638683",
    "ASIST.CAMBIO CARRIL:D+M(TRAB.ADIC.)",
    "BALANCEO 260415174154788",
    "BIELETA I.ESTABI.:D+M",
    "BRAZO SUSPENS.DL.INF.:D+M (LLANTA DESM.)",
    "CALIPER FRENO DL.I.:SOLT-FIJ.",
    "CAMARA 360 GRADOS: AJUSTAR",
    "CARGA DE GAS 260505174818841",
    "CARGA DE GAS 260512135935854",
    "CARGA GAS 260327180816370",
    "DEPOS.EXPANSION:D+M",
    "DES-/MONTAR FLUIDO AIRE ACONDIC.",
    "DES+MON RIN TRA.D.",
    "ESCANEO AIRBAG 260512135935902",
    "FILTRO AIRE:D+M",
    "FLECHA CARDAN DIREC.I.:D+M",
    "FLECHA MOTRIZ DL.I.CPL.:D+M",
    "GALON ANTICONGELANTE 260327180816379",
    "KA) AMORTIG./MANGUETA DL.D.:SOLT-FIJ",
    "KA) AMORTIG.DL.D.:D+M",
    "KA) AMORTIG.DL.D./CARROCER.:SOLT.-FIJ",
    "KA) AMORTIGUADOR/S DL.:D+M TRAB.ADIC.",
    "KA) LLANTA/LLANTAS:D+M TRAB.ADIC.",
    "KA) RIN DL.I.:D+M",
    "LIQUIDO REFRIGERANTE DES+MON/SUSTITUIR",
    "LIQUIDO REFRIGERANTE REPARAR",
    "LLANTA DL.D.:D+M(LLANTA DESMONT.)",
    "LLANTA DL.D.:MONTAR/BALANCEAR",
    "LLANTA Y/O RIN(ES):D+M",
    "PROG SENSOR RADAR 260327180816362",
    "R RIN DEL DER 260525143507914",
    "R RIN DEL IZQ 260511145013209",
    "RADIADOR EGR:D+M",
    "REC.RODAM.DL.D.:D+M",
    "REFRIGERANTE A.ACOND REPARAR",
    "REP ARNES SENSO REVE 260609083801049",
    "RESONADOR INF. REPARAR",
    "RIN TRA.D. REPARAR",
    "ROTULA BIELETA DIREC.I.:D+M",
    "ROTULA SOP.DL.I.:SOLT-FIJ",
    "SENSOR BOLSA AIRE DL.:D+M",
    "SENSOR TR.CN.ASIST.ESTAC.:D+M",
    "SOP.D.RADIADOR:SUST.",
    "VA) BRAZO SUSPENS.DL.D.:D+M",
    "VA) BRAZO SUSPENS.DL.I./D.:D+M TRAB.ADIC.",
    "VA) LLANTA DL.D.:BALANCEAR",
    "VA) PUNTA BIELETA DIREC.:D+M TRAB.ADIC.",
    "VA) ROTULA BIELETA DIREC.D.:D+M",
    "VALV.PRESION DEL.I. REPARAR",
}

def clasificar_categoria(categoria: str, descripcion: str) -> str:
    """Si la partida es de Hojalatería y su descripción está en la lista, → Mecánica."""
    if categoria == "HOJALATERIA" and descripcion.strip() in CONCEPTOS_MECANICA:
        return "MECANICA"
    return categoria

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_ot(filename: str) -> str:
    return os.path.splitext(filename)[0]

def es_no_parte_refaccion(token: str) -> bool:
    return not re.match(r'^\d+\.\d+$', token)

def parse_lines(lines, seccion_inicial=None):
    pat_refac = re.compile(r'^(.+?)\s+(\S+)\s+\$\s*([\d,]+\.\d{2})\s*$')
    pat_ut    = re.compile(r'^(.+?)\s+(\d+\.\d+)\s+\$\s*([\d,]+\.\d{2})\s*$')
    pat_tpp   = re.compile(r'^(.+?)\s+TPP\s+(\d+\.\d+)\s+\$\s*([\d,]+\.\d{2})\s*$')

    partidas = []
    seccion  = seccion_inicial

    for line in lines:
        ls = line.strip()
        if not ls:
            continue

        if re.match(r'^REFACCIONES\b', ls):
            seccion = "REFACCIONES"; continue
        if re.match(r'^PINTURA\b', ls):
            seccion = "PINTURA"; continue
        if re.match(r'^MANO DE OBRA\b', ls, re.IGNORECASE):
            seccion = "HOJALATERIA"; continue

        if seccion is None:
            continue

        if re.search(
            r'DESCRIPCION|NO\. PARTE|^Subtotal|^IVA|^Total\b|No Efectivo|^UT\s|R E S U M E N|SUMA TOTAL|DEDUCIBLE|DEMÉRITO',
            ls, re.IGNORECASE
        ):
            continue

        if seccion == "REFACCIONES":
            m = pat_refac.match(ls)
            if m and es_no_parte_refaccion(m.group(2)):
                desc = m.group(1).strip()
                partidas.append({
                    "CATEGORIA":      clasificar_categoria("REFACCIONES", desc),
                    "DESCRIPCION":    desc,
                    "NO. PARTE / UT": m.group(2),
                    "MONTO":          float(m.group(3).replace(",", "")),
                })
        else:
            m = pat_tpp.match(ls)
            if m:
                desc = m.group(1).strip()
                partidas.append({
                    "CATEGORIA":      clasificar_categoria(seccion, desc),
                    "DESCRIPCION":    desc,
                    "NO. PARTE / UT": m.group(2),
                    "MONTO":          float(m.group(3).replace(",", "")),
                })
                continue
            m = pat_ut.match(ls)
            if m:
                desc = m.group(1).strip()
                partidas.append({
                    "CATEGORIA":      clasificar_categoria(seccion, desc),
                    "DESCRIPCION":    desc,
                    "NO. PARTE / UT": m.group(2),
                    "MONTO":          float(m.group(3).replace(",", "")),
                })

    return partidas, seccion


def extract_all(pdf_bytes: bytes, ot: str) -> list[dict]:
    partidas = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            pw, ph = page.width, page.height

            left_text  = page.crop((0, 0, pw * 0.50, ph)).extract_text() or ""
            left_rows, last_seccion = parse_lines(left_text.split("\n"))
            partidas.extend(left_rows)

            right_text  = page.crop((pw * 0.50, 0, pw, ph)).extract_text() or ""
            right_lines = right_text.split("\n")

            tiene_encabezado = any(
                re.match(r'^(REFACCIONES|PINTURA|MANO DE OBRA)\b', l.strip(), re.IGNORECASE)
                for l in right_lines
            )
            seccion_der = None if tiene_encabezado else last_seccion
            right_rows, _ = parse_lines(right_lines, seccion_inicial=seccion_der)
            partidas.extend(right_rows)

    seen = set()
    result = []
    for p in partidas:
        key = (p["CATEGORIA"], p["DESCRIPCION"], p["MONTO"])
        if key not in seen:
            seen.add(key)
            result.append({"OT": ot, **p})

    return result


# ── Excel ─────────────────────────────────────────────────────────────────────

CAT_COLORS = {
    "REFACCIONES": "1A3A5C",
    "PINTURA":     "C8392B",
    "HOJALATERIA": "2E7D32",
    "MECANICA":    "6A1B9A",
}

def build_excel(all_rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuación"

    AZUL_OSC   = "1A3A5C"
    AZUL_MED   = "2E6DA4"
    GRIS_CLARO = "F0F4F8"
    BLANCO     = "FFFFFF"

    thin  = Side(style="thin", color="CCCCCC")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "VALUACIÓN QUÁLITAS – REFACCIONES / PINTURA / HOJALATERÍA / MECÁNICA"
    c.font = Font(name="Arial", bold=True, size=13, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL_OSC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, hdr in enumerate(["OT", "CATEGORÍA", "DESCRIPCIÓN", "NO. PARTE / UT", "MONTO"], 1):
        c = ws.cell(2, col, hdr)
        c.font = Font(name="Arial", bold=True, size=10, color=BLANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde
    ws.row_dimensions[2].height = 20

    # Orden de categorías en el Excel
    CAT_ORDER = {"REFACCIONES": 0, "PINTURA": 1, "HOJALATERIA": 2, "MECANICA": 3}
    sorted_rows = sorted(all_rows, key=lambda r: (r["OT"], CAT_ORDER.get(r["CATEGORIA"], 9)))
    current_row = 3
    data_start  = 3

    for ot_key, ot_group in groupby(sorted_rows, key=lambda r: r["OT"]):
        ot_list  = list(ot_group)
        ot_first = current_row

        for cat_key, cat_group in groupby(ot_list, key=lambda r: r["CATEGORIA"]):
            cat_list  = list(cat_group)
            cat_first = current_row
            cat_color = CAT_COLORS.get(cat_key, AZUL_MED)

            for i, row in enumerate(cat_list):
                bg   = BLANCO if i % 2 == 0 else GRIS_CLARO
                fill = PatternFill("solid", fgColor=bg)
                fnt  = Font(name="Arial", size=10)

                for col in range(1, 6):
                    c = ws.cell(current_row, col)
                    c.fill = fill; c.font = fnt; c.border = borde

                ws.cell(current_row, 1).value = row["OT"]
                ws.cell(current_row, 1).alignment = Alignment(horizontal="center", vertical="center")

                c2 = ws.cell(current_row, 2)
                c2.value = row["CATEGORIA"]
                c2.font  = Font(name="Arial", size=10, bold=True, color=cat_color)
                c2.alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(current_row, 3).value = row["DESCRIPCION"]
                ws.cell(current_row, 3).alignment = Alignment(vertical="center")
                ws.cell(current_row, 4).value = row["NO. PARTE / UT"]
                ws.cell(current_row, 4).alignment = Alignment(horizontal="center", vertical="center")

                c5 = ws.cell(current_row, 5)
                c5.value = row["MONTO"]
                c5.alignment = Alignment(horizontal="right", vertical="center")
                c5.number_format = '"$"#,##0.00'
                current_row += 1

            cat_last = current_row - 1

            ws.merge_cells(f"A{current_row}:D{current_row}")
            sl = ws.cell(current_row, 1)
            sl.value = f"Subtotal {cat_key} — OT {ot_key}"
            sl.font  = Font(name="Arial", bold=True, size=10, color=BLANCO)
            sl.fill  = PatternFill("solid", fgColor=cat_color)
            sl.alignment = Alignment(horizontal="right", vertical="center")

            sv = ws.cell(current_row, 5)
            sv.value = f"=SUM(E{cat_first}:E{cat_last})"
            sv.font  = Font(name="Arial", bold=True, size=10, color=BLANCO)
            sv.fill  = PatternFill("solid", fgColor=cat_color)
            sv.alignment = Alignment(horizontal="right", vertical="center")
            sv.number_format = '"$"#,##0.00'
            sv.border = borde
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        ot_last = current_row - 1

        ws.merge_cells(f"A{current_row}:D{current_row}")
        tl = ws.cell(current_row, 1)
        tl.value = f"TOTAL OT {ot_key}"
        tl.font  = Font(name="Arial", bold=True, size=11, color=BLANCO)
        tl.fill  = PatternFill("solid", fgColor=AZUL_OSC)
        tl.alignment = Alignment(horizontal="right", vertical="center")

        tv = ws.cell(current_row, 5)
        tv.value = f"=SUM(E{ot_first}:E{ot_last})"
        tv.font  = Font(name="Arial", bold=True, size=11, color=BLANCO)
        tv.fill  = PatternFill("solid", fgColor=AZUL_OSC)
        tv.alignment = Alignment(horizontal="right", vertical="center")
        tv.number_format = '"$"#,##0.00'
        tv.border = borde
        ws.row_dimensions[current_row].height = 22
        current_row += 2

    ws.merge_cells(f"A{current_row}:D{current_row}")
    gl = ws.cell(current_row, 1)
    gl.value = "TOTAL GENERAL"
    gl.font  = Font(name="Arial", bold=True, size=12, color=BLANCO)
    gl.fill  = PatternFill("solid", fgColor="C8392B")
    gl.alignment = Alignment(horizontal="right", vertical="center")

    gv = ws.cell(current_row, 5)
    gv.value = f"=SUM(E{data_start}:E{current_row - 1})"
    gv.font  = Font(name="Arial", bold=True, size=12, color=BLANCO)
    gv.fill  = PatternFill("solid", fgColor="C8392B")
    gv.alignment = Alignment(horizontal="right", vertical="center")
    gv.number_format = '"$"#,##0.00'
    gv.border = borde
    ws.row_dimensions[current_row].height = 24

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ────────────────────────────────────────────────────────────────────────

uploaded_files = st.file_uploader(
    "Sube uno o más PDFs de valuación Quálitas",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

if uploaded_files:
    st.markdown(f"**{len(uploaded_files)} archivo(s) cargado(s):**")
    for f in uploaded_files:
        st.markdown(f"- `{f.name}`")

    if st.button("⚙️  Extraer y generar Excel"):
        all_rows = []
        errores  = []
        progress = st.progress(0)

        for i, f in enumerate(uploaded_files):
            ot = get_ot(f.name)
            try:
                rows = extract_all(f.read(), ot)
                if rows:
                    all_rows.extend(rows)
                else:
                    errores.append(f"⚠️ `{f.name}` — no se encontraron partidas.")
            except Exception as e:
                errores.append(f"❌ `{f.name}` — error: {e}")
            progress.progress((i + 1) / len(uploaded_files))

        for msg in errores:
            st.warning(msg)

        if all_rows:
            excel_bytes = build_excel(all_rows)
            total = sum(r["MONTO"] for r in all_rows)

            cats = {}
            for r in all_rows:
                cats[r["CATEGORIA"]] = cats.get(r["CATEGORIA"], 0) + r["MONTO"]

            st.markdown('<div class="result-box">', unsafe_allow_html=True)
            st.markdown(f"✅ **{len(all_rows)} partidas** de **{len(uploaded_files)} PDF(s)**")
            for cat in ["REFACCIONES", "PINTURA", "HOJALATERIA", "MECANICA"]:
                if cat in cats:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;• **{cat}**: ${cats[cat]:,.2f}")
            st.markdown(f"💰 **Total general: ${total:,.2f}**")
            st.markdown('</div>', unsafe_allow_html=True)

            df_show = pd.DataFrame(all_rows).copy()
            df_show["MONTO"] = df_show["MONTO"].map(lambda x: f"${x:,.2f}")
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            out_name = (
                f"{get_ot(uploaded_files[0].name)}_valuacion.xlsx"
                if len(uploaded_files) == 1
                else "valuacion_qualitas.xlsx"
            )

            st.download_button(
                label="📥  Descargar Excel",
                data=excel_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.error("No se pudieron extraer partidas de ningún PDF.")
else:
    st.info("👆 Sube los PDFs de valuación para comenzar.")

st.markdown("---")
st.caption("Extractor Valuaciones Quálitas · Refacciones · Pintura · Hojalatería · Mecánica")
